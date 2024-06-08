"""
A one-off script to combine multiple excel files into
a single excel file with multiple sheets
with formatting properties copied over.
"""

import os
from openpyxl import Workbook, load_workbook
from copy import copy

def copy_format(source_cell, target_cell):
    """
    Copy formatting properties from a source cell to a target cell.
    """
    target_cell.font = copy(source_cell.font)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)

    if source_cell.hyperlink:
        target_cell._hyperlink = source_cell.hyperlink


def combine_sheets_keeping_format(input_folder: str, output_file: str):
    """
    Combines all excel files in the input folder
    into a single excel file with multiple sheets
    with copied formats

    PARAMETERS:
        input_folder: folder containing the excel files
        output_file: output excel file

    RETURNS:
        None
    """
    wb = Workbook()

    for file in os.listdir(input_folder):
        if file.endswith(".xlsx"):
            wb_temp = load_workbook(os.path.join(input_folder, file))
            sheet = wb_temp.active
            new_sheet = wb.create_sheet(title=file.split(".")[0])

            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    copy_format(cell, new_cell)

    wb.save(output_file)

combine_sheets_keeping_format("data/example_data", "data/combined_sheets.xlsx")
