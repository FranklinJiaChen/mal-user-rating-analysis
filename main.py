"""
This script fetches the anime list of a user from MyAnimeList
and creates an Excel file
Rows: Rated anime in the user's list
Columns: Title, User Rating, MAL Score
"""
import os
import requests
from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from anime_data_model import UserAnime
from util import try_with_exponential_backoff as try_with_backoff

load_dotenv()
CLIENT_ID = os.getenv("CLIENT_ID")
FIELDS = "mean,num_favorites,statistics,source,related_anime,media_type"


def get_user_anime(name, offset=0) -> list[tuple[int, int]]:
    """
    Fetches up to 100 anime from the list of a user from MyAnimeList

    PARAMETERS:
        name: The username of the user
        offset: The offset to fetch the anime from

    RETURNS:
        A list of rated anime by the user in the format
        (anime id, score given by user)

    RAISES:
        HTTPError: If the request to the API fails
    """
    url = (
        f'https://api.myanimelist.net/v2/users/{name}/animelist'
        f'?fields=list_status&limit=100&offset={offset}'
    )

    try:
        response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
        response.raise_for_status()
        animes = response.json()['data']

        anime_list: list[tuple[int, int]] = []
        for anime in animes:
            if (anime['list_status']['score']):
                anime_list.append((anime['node']['id'], anime['list_status']['score']))
        response.close()

        return anime_list

    except requests.exceptions.HTTPError as e:
        raise e


def get_all_user_anime(name: str) -> list[tuple[int, int]]:
    """
    Retrieves a list of rated anime by the user.

    PARAMETERS:
        name: The username of the user.
        offset: The offset to fetch the anime from.

    RETURNS:
        A list of rated anime by the user in the format
        (anime id, score given by user)
    """
    offset = 0
    anime_list = []
    while 1:
        new_animes = try_with_backoff(get_user_anime, name, offset)
        if not new_animes:
            break
        anime_list.extend(new_animes)
        offset += 100
    return anime_list


def fetch_anime_data(anime_id, field: str) -> dict:
    """
    Fetches data for a specific anime from MyAnimeList

    PARAMETERS:
        anime_id: The id of the anime
        field: The fields to fetch

    RETURNS:
        A dictionary containing the anime data in json format

    RAISES:
        HTTPError: If the request to the API fails
    """
    url = f'https://api.myanimelist.net/v2/anime/{anime_id}?fields={field}'
    try:
        response = requests.get(url, headers={'X-MAL-CLIENT-ID': CLIENT_ID})
        response.raise_for_status()
        anime_data = response.json()
        response.close()
        return anime_data
    except requests.exceptions.HTTPError as e:
        raise e


def create_anime(anime: tuple[int, int]) -> UserAnime:
    """
    Creates an anime object from the data fetched from MyAnimeList

    PARAMETERS:
        anime: A tuple containing the anime id and the score given by the user

    RETURNS:
        A UserAnime object of the anime
    """
    anime_id, user_score = anime
    anime_data = try_with_backoff(fetch_anime_data, anime_id, FIELDS)

    anime = UserAnime(anime_data, user_score)
    return anime


def create_sheet(workbook, anime_list) -> None:
    """
    Creates an Excel sheet with the anime data

    PARAMETERS:
        workbook: The workbook to create the sheet in
        anime_list: A list of UserAnime objects
    """
    TITLE = 1
    USER_RATING = 2
    MAL_SCORE = 3

    N = 5
    CORRELATION = 6

    headers = ["Title", "User Rating", "MAL Score", "", "n", "Correlation"]
    sheet = workbook.active
    sheet.title = "Anime List"

    for i, header in enumerate(headers):
        sheet.cell(1, i+1).value = header
        sheet.cell(1, i+1).alignment = Alignment(horizontal='center')
        sheet.cell(1, i+1).font = Font(bold=True)
        sheet.freeze_panes = 'B2'

    for j, anime in enumerate(anime_list):
        sheet.cell(j+2, TITLE).value = anime.title
        sheet.cell(j+2, USER_RATING).value = anime.user_rating
        sheet.cell(j+2, MAL_SCORE).value = anime.rating

        sheet.cell(j+2, TITLE).hyperlink = anime.get_mal_link()

        sheet.cell(j + 2, USER_RATING).alignment = Alignment(
            horizontal='center'
        )
        sheet.cell(j + 2, MAL_SCORE).alignment = Alignment(
            horizontal='center'
        )

    sheet.cell(2, N).value = len(anime_list)
    sheet.cell(2, CORRELATION).value = f'=correl(B:B, C:C)'

def main() -> None:
    """
    Main function to fetch the anime list of a user and create an Excel file
    """
    user = ""  # Add the MAL username of the user here

    user_animes = get_all_user_anime(user)
    anime_list = [print(f"Progress: {i + 1}/{len(user_animes)}") or
                  create_anime(anime) for i, anime in enumerate(user_animes)]

    workbook = Workbook()
    create_sheet(workbook, anime_list)
    workbook.save(f"data/{user}_anime_list.xlsx")


if __name__ == "__main__":
    main()
